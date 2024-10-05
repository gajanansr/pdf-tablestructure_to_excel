import fitz  # PyMuPDF
import cv2
import pytesseract
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Border, Side
import matplotlib.pyplot as plt

# Function to extract the last page from a PDF
def extract_last_page(pdf_path, output_image_path):
    pdf_document = fitz.open(pdf_path)
    last_page = pdf_document.page_count - 1
    page = pdf_document.load_page(last_page)
    pix = page.get_pixmap()
    pix.save(output_image_path)

# Function to detect table structure using OpenCV
def detect_table_structure(image_path):
    image = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
    _, binary = cv2.threshold(image, 150, 255, cv2.THRESH_BINARY_INV)

    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 15))
    vertical_lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, vertical_kernel, iterations=2)

    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (15, 1))
    horizontal_lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, horizontal_kernel, iterations=2)

    table_structure = cv2.addWeighted(vertical_lines, 0.5, horizontal_lines, 0.5, 0.0)

    contours, _ = cv2.findContours(table_structure, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

    # Optional: Draw contours for debugging
    debug_image = cv2.cvtColor(image, cv2.COLOR_GRAY2BGR)
    for contour in contours:
        if cv2.contourArea(contour) > 100:  # Filter out small contours
            x, y, w, h = cv2.boundingRect(contour)
            cv2.rectangle(debug_image, (x, y), (x + w, y + h), (0, 255, 0), 2)  # Draw rectangle

    cv2.imwrite('debug_table_structure.png', debug_image)  # Save debug image for inspection

    return contours

# Function to extract cell text using Tesseract
def extract_cell_text(image_path, contours):
    image = cv2.imread(image_path)
    cell_data = []

    for contour in contours:
        if cv2.contourArea(contour) > 100:  # Filter out small contours
            x, y, w, h = cv2.boundingRect(contour)
            cell_image = image[y:y + h, x:x + w]
            text = pytesseract.image_to_string(cell_image, config='--psm 6').strip()
            cell_data.append((x, y, text))

    # Debugging: Print the extracted cell data
    print("Extracted Cell Data:")
    for x, y, text in cell_data:
        print(f"Cell at ({x}, {y}): {text}")

    return cell_data

# Function to create the Excel sheet with merged cells
def create_excel_with_borders(cell_data, output_excel_path):
    wb = Workbook()
    ws = wb.active

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Create a dictionary to track merged cells
    merged_cells = {}

    # Sort cell data by (y, x) coordinates to ensure rows are processed in order
    cell_data.sort(key=lambda cell: (cell[1], cell[0]))

    # Process cell data to merge cells based on coordinates
    for x, y, text in cell_data:
        row = int(y / 20) + 1
        col = int(x / 50) + 1
        ws.cell(row=row, column=col, value=text).border = thin_border

        # Debugging: Log each cell position and its text
        print(f"Placing text '{text}' at Excel cell ({row}, {col})")

        # Check if the current cell should merge with others
        if (row, col) in merged_cells:
            merged_cells[(row, col)].append((x, y, text))
        else:
            merged_cells[(row, col)] = [(x, y, text)]

    # Merging cells based on proximity and content
    for (row, col), texts in merged_cells.items():
        if len(texts) > 1:
            x_positions = [t[0] for t in texts]
            if max(x_positions) - min(x_positions) < 50:  # Merge if close in x-direction
                ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + len(texts) - 1)
                ws.cell(row=row, column=col).value = ' '.join(t[2] for t in texts)
                print(f"Merging cells in row {row} from column {col} to {col + len(texts) - 1}")

    # Save the Excel file
    wb.save(output_excel_path)

# Function to create and save a truth table
def create_truth_table(cell_data):
    truth_table = []
    rows, cols = 7, 20

    for i in range(rows):
        row_data = [text for _, _, text in cell_data[i * cols: (i + 1) * cols]]
        while len(row_data) < cols:
            row_data.append('')
        truth_table.append(row_data)

    fig, ax = plt.subplots()
    ax.axis('tight')
    ax.axis('off')
    table = ax.table(cellText=truth_table, loc='center', cellLoc='center')
    table.auto_set_font_size(False)
    table.set_fontsize(8)
    table.scale(1.2, 1.2)

    plt.savefig('truth_table.png', bbox_inches='tight', dpi=300)

# Main function to process PDF and create Excel
def process_pdf_to_excel(pdf_path, output_image_path, output_excel_path):
    extract_last_page(pdf_path, output_image_path)
    contours = detect_table_structure(output_image_path)
    cell_data = extract_cell_text(output_image_path, contours)
    create_excel_with_borders(cell_data, output_excel_path)
    create_truth_table(cell_data)

# Example usage:
pdf_path = "./Alibag.pdf"
output_image_path = "last_page_image.png"
output_excel_path = "output_table_structure.xlsx"
process_pdf_to_excel(pdf_path, output_image_path, output_excel_path)
